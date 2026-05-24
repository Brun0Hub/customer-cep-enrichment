import time
import pandas as pd
from openpyxl import load_workbook

from services.cep_service import get_cep_data
from services.excel_service import load_excel, ensure_columns


def enrich_customers(input_file):
    df = load_excel(input_file)
    df = ensure_columns(df)

  
    wb = load_workbook(input_file)
    ws = wb.active

    for i, row in df.iterrows():
        cep = row["Cep"]

     
        if pd.notna(row.get("Address")) and row["Address"] != "":
            continue

        data = get_cep_data(cep)

        if data:
            df.at[i, "Address"] = data.get("logradouro", "")
            df.at[i, "Neighborhood"] = data.get("bairro", "")
            df.at[i, "City"] = data.get("localidade", "")
            df.at[i, "State"] = data.get("uf", "")
            df.at[i, "Region"] = data.get("regiao", "")

           
            excel_row = i + 2

            ws[f"F{excel_row}"] = df.at[i, "Address"]
            ws[f"I{excel_row}"] = df.at[i, "Neighborhood"]
            ws[f"J{excel_row}"] = df.at[i, "City"]
            ws[f"K{excel_row}"] = df.at[i, "State"]
            ws[f"L{excel_row}"] = df.at[i, "Region"]

        time.sleep(0.3)


    wb.save(input_file)

    print("✅ Planilha atualizada com sucesso!")